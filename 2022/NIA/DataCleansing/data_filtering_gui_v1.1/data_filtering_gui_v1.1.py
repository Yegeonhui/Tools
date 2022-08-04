import PySimpleGUI as sg
from data_filtering import videofiltering, imagefiltering
import os
from glob import glob
import openpyxl

class MakeGUI():
    def makegui(self):
        sg.theme('Black')
        layout = [
                  [sg.Text('Input Directory',size=(20, 1))],
                  [sg.InputText('mp4/jpg file Path', key='Path'), sg.FolderBrowse('Select', size=(10, 1))],
                  [sg.Text('Sampling Intervals',size=(20, 1))],
                  [sg.InputText('Sampling Intervals(seconds)', key='intervals')],
                  [sg.Text('Input Data Type'), sg.Checkbox('mp4', key="mp4"), sg.Checkbox('jpg', key="jpg"), sg.Button('Run', size=(11, 1), key='Run')],
                  [sg.ProgressBar(1, orientation='h', size=(40,20), key='progress')]
                  ]
        window = sg.Window('Data filtering', layout, grab_anywhere = True).Finalize()
        return window


def getwaterinfo(info_path):
    wb = openpyxl.load_workbook(info_path)

    # 수질환경정보
    ws1 = wb['Sheet1']
    water_env = {}
    for c in range(1, 9):
        if ws1.cell(2,c).value == None:
            return False, water_env
        else:
            water_env[ws1.cell(1,c).value] = ws1.cell(2,c).value
    return True, water_env


def getobjectinfo(info_path, imagename):
    wb = openpyxl.load_workbook(info_path) 
    ws2 = wb['Sheet2']
    obj_info = []
    for r2 in range(2, ws2.max_row):
        info = []
        for c2 in range(3, 6):
            if ws2.cell(r2, 1).value == imagename:
                info.append(ws2.cell(r2, c2).value)
        if len(info) != 0:
            obj_info.append(info)
    return obj_info


def checkwaterinfo(water_env):
    flag_array = [True for i in range(1, 9)]
    if 0 > water_env['Temp'] or water_env['Temp'] > 40:
        flag_array[0] = False
    if 0 > water_env['Salinity'] or water_env['Salinity'] > 40:
        flag_array[1] = False
    if 0 > water_env['Do'] or water_env['Do'] > 15:
        flag_array[2] = False
    if 6 > water_env['pH'] or water_env['pH'] > 9:
        flag_array[3] = False
    if 0 > water_env['Transparency'] or water_env['Transparency'] > 15:
        flag_array[4] = False
    if 33.11 > water_env['lon'] or water_env['lon'] > 38.61:
        flag_array[5] = False
    if 124.6 > water_env['lat'] or water_env['lat'] > 131.87:
        flag_array[6] = False
    if 0 > water_env['Depth']:
        flag_array[7] = False
    return flag_array


m = MakeGUI()
window = m.makegui()

while True:    
    event, values = window.read()     
    progress_bar = window.FindElement('progress')
    if event == 'Run':
        path = values['Path']
        info_path = path + '/information.xlsx'
        flag, water_env = getwaterinfo(info_path)
        
        if not flag:
            sg.Popup('수질환경정보 null값 확인', keep_on_top=True)
            continue
            
        flag_array = checkwaterinfo(water_env)
        if False in flag_array:
            sg.Popup('수질환경정보 이상값 확인', keep_on_top=True)
            continue

        if values['mp4']:
            videolist = glob(path + "/*.mp4")
            interval = int(values['intervals'])
            cnt = 1
            for video in videolist:
                progress_bar.UpdateBar(cnt, len(videolist))
                videoname = os.path.split(video)[-1]
                print(videoname)
                savepath = path + '/' + os.path.splitext(videoname)[0]
                os.makedirs(savepath, exist_ok=True)
                videofiltering(video, interval, savepath, water_env)

                cnt += 1

        if values['jpg']:
            imagelist = glob(path + "/*.jpg")
            os.makedirs(path + '/filtering', exist_ok=True)
            cnt = 1
            for image in imagelist:
                #2022_03_21_0001.jpg
                progress_bar.UpdateBar(cnt, len(imagelist))
                imagenamejpg = os.path.split(image)[-1]
                print(imagenamejpg)
                #2022_03_21_0001
                imagename = os.path.splitext(imagenamejpg)[0]

                savepath = path + '/filtering'
                obj_info = getobjectinfo(info_path, imagename) 
                
                imagefiltering(image, savepath, water_env, obj_info)
                
                cnt += 1
                
    if event in (None, 'Exit'):
        break
